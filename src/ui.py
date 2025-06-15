from tkinter import Text
import webbrowser
from tkinter import Tk, StringVar, IntVar, Label, Entry, Button, END
from tkinter.ttk import Combobox
from tkinter.filedialog import askopenfilename
from tkinter.messagebox import showwarning, showerror, showinfo
from main import read_bms_config, main_prase
from utils import get_text_encoding
import json
import re
import os
import time
import csv
import copy
import openpyxl
import sys
import win32com.client



version = '<version>'
win_width = 0
win_height = 0
file_path = ''
unvalid_header = []


def creat_window():
    win_width = 500 #int(root.winfo_screenwidth() / 3)
    win_height = 256 #int(root.winfo_screenheight() / 3)
    root.title('BMS翻译官 ' + version)
    root.geometry(f'{win_width}x{win_height}+{int(root.winfo_screenwidth()/3)}+{int(root.winfo_screenheight()/3)}')
    return [win_width, win_height]

def save_config():
    config['timestamp'] = int(time.time())
    config['id_place'] = id_place.get()
    config['data_place'] = data_place.get()
    config['split'] = split_entry.get()
    config['valid_row'] = valid_entry.get()
    config['protocols_type'] = protocols_entry.get()
    with open('./config/config', 'w', encoding='utf-8') as fp:
        json.dump(config, fp)

def creat_entry():
    # 帧ID位置 的选项
    id_label = Label(root, text='帧ID所在的列', font=('黑体', 11))
    id_label.place(relx=0.09, y=25)
    var_id = StringVar()
    even_numbers_1 = [f"{i+1} | {chr(65+i)}" for i in range(9)]
    frame_id_place = Combobox(root, textvariable=var_id, values=even_numbers_1)
    frame_id_place.place(relx=0.1, y=50, width=110)
    frame_id_place.set('1')

    # 帧数据位置的选项
    data_label = Label(root, text='帧数据所在的列', font=('黑体', 11))
    data_label.place(relx=0.35, y=25)
    var_data = StringVar()
    even_numbers_2 = [f"{i+1} | {chr(65+i)}" for i in range(9)]
    frame_date_place = Combobox(
        root, textvariable=var_data, values=even_numbers_2)
    frame_date_place.place(relx=0.35, y=50, width=120)

    # 数据分隔符
    split_label = Label(root, text='数据分隔符', font=('黑体', 11))
    split_label.place(relx=0.62, y=25)
    var_split = StringVar()
    split_entry = Combobox(root, textvariable=var_split,
                           values=['tab(\\t)', '英文逗号(,)', '英文分号(;)', '竖线(|)'])
    split_entry.place(relx=0.62, y=50, width=110)

    # 选择有效数据行
    valid_num_label = Label(root, text='有效数据开始行', font=('黑体', 11))
    valid_num_label.place(relx=0.09, y=90)
    var_valid = IntVar()
    even_numbers_3 = [str(x) for x in range(1, 5)]
    valid_num_entry = Combobox(
        root, textvariable=var_valid, values=even_numbers_3)
    valid_num_entry.place(relx=0.1, y=90+25, width=110)

    # 选择解析协议
    global bms_config
    protocols_label = Label(root, text='选择BMS协议', font=('黑体', 11))
    protocols_label.place(relx=0.62, y=90)
    protocols_var_valid = StringVar()
    even_numbers_4 = list(bms_config.keys()) # type: ignore
    protocols_num_entry = Combobox(
        root, textvariable=protocols_var_valid, values=even_numbers_4)
    protocols_num_entry.place(relx=0.62, y=90+25, width=110)

    # 底部信息
    def open_url(event):
        webbrowser.open_new("https://github.com/liuyu80/bms-translator/releases/latest")

    git_url = Text(root, height=1, width=60, wrap="none", bg=root.cget('bg'), fg='blue', cursor="hand2", borderwidth=0, highlightthickness=0)
    git_url.insert(1.0, '下载最新软件')
    git_url.config(state='disabled')
    git_url.bind("<Button-1>", open_url)
    git_url.place(relx=0.75, rely=0.9)

    # 文件路径显示
    file_path_entry = Entry(root)
    file_path_entry.place(relx=0.1, y=160, width=(int(win_width)/3) * 2 + 35)
    file_path_entry.insert(0, " 选择文件路径")
    return [frame_id_place, frame_date_place, split_entry, valid_num_entry, file_path_entry, protocols_num_entry]


def ui_path_check(path):
    path = file_path_entry.get()
    if path == '文件路径' or path == '':
        showwarning('警告', '路径为空')
        return 0
    file_path, file_name = os.path.split(path)
    name, style = os.path.splitext(file_name)
    # if style not in ('.CSV', '.csv', '.xls', '.XLS'):
    #     showwarning('警告', '文件名不正确')
    #     return 0
    # if name[-2:] == '-译':
    #     showwarning('警告', '该文件已翻译, 请选择要翻译的报文文件')
    #     return 0
    return 1


def open_file_manager():
    global file_path
    file_path = askopenfilename(filetypes=[("支持的文件", "*.csv *.xls *.xlsx *.asc"), ("所有文件", "*.*")])
    file_path_entry.delete(0, END)
    file_path_entry.insert(0, file_path)

# 
def read_csv(path, split_s, valid_num):
    with open(path, "r", encoding=get_text_encoding(path)) as file:
        csv_reader = csv.reader(file, delimiter=split_s)
        data = []
        for line in csv_reader:
            data.append(line)
    global unvalid_header
    for line in data[: valid_num-1]:
        unvalid_header.append([line])
    unvalid_header.append([''])

    return data[valid_num-1:]

def read_excel(path, valid_num):
    try:
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        if sheet is None:
            showerror('错误', 'Excel文件没有活动工作表或文件为空。')
            return []
        data = []
        for row in sheet.iter_rows():
            row_data = [cell.value for cell in row]
            data.append(row_data)
        
        global unvalid_header
        for line in data[: valid_num-1]:
            unvalid_header.append([line])
        unvalid_header.append([''])
        
        return data[valid_num-1:]
    except Exception as e:
        showerror('错误', f'读取Excel文件失败: {e}')
        return []

def creat_csv(path, data_df, splite_s):
    file_path, file_name = os.path.split(path)  # 路径切割, 得到路径和文件名
    # 生成新的文件 保存解析后的数据
    base_name = os.path.splitext(file_name)[0]
    csv_name = f"{base_name}-译.csv"
    with open(os.path.join(file_path, csv_name), 'w', newline='', encoding='utf-8') as file:
        # 使用制表符作为分隔符创建 CSV 的 writer 对象
        writer = csv.writer(file, delimiter=splite_s)
        # 写入数据行
        for row in data_df:
            writer.writerow(row)
    os.startfile(os.path.join(file_path, csv_name))  # 自动使用系统默认应用打开该文件

def creat_excel(path, data_df):
    file_path, file_name = os.path.split(path)
    base_name = os.path.splitext(file_name)[0]
    excel_name = f"{base_name}-译.xlsx"
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    if sheet is None:
        showerror('错误', 'Excel文件没有活动工作表或文件为空。')
        return
    
    for row_data in data_df:
        sheet.append(row_data)
        
    output_path = os.path.join(file_path, excel_name)
    try:
        workbook.save(output_path)
        os.startfile(output_path)
    except Exception as e:
        showerror('错误', f'保存Excel文件失败: {e}')

def check_data(data: list, idx: int , datax: int) -> bool:
    if not data or len(data) < 2: # Handle empty or too small data
        showerror('错误', '文件数据为空或数据量不足，请检查文件内容或分隔符')
        return False
    
    id_idx = idx - 1
    if id_idx < 0 or id_idx >= len(data):
        showerror('错误', '帧ID列超出数据范围，请选择正确的帧ID列')
        return False
    if not data[id_idx] or len(data[id_idx]) < 8:
        showerror('错误', '帧ID列数据格式不正确或长度不足，请选择正确的帧ID列')
        return False
    
    data_idx = datax - 1
    if data_idx < 0 or data_idx >= len(data):
        showerror('错误', '帧数据列超出数据范围，请选择正确的帧数据列')
        return False
    if not data[data_idx] or len(data[data_idx]) < 2:
        showerror('错误', '帧数据列数据格式不正确或长度不足，请选择正确的帧数据列')
        return False
    
    return True


def parse_asc_file(asc_file) -> list:
    regex_pattern = r"^(\d+\.\d+).*?(\w+f456|\w+56f4).*?d\s(\d+)\s(.*?)$"
    output_csv_path = 'matched_data.csv'
    matched_data_list = []

    with open(asc_file, 'r', encoding='utf-8') as infile:
        for line_num, line in enumerate(infile, 1):
            line = line.strip() # 移除行末的空白符，包括换行符
            match = re.match(regex_pattern, line)
            if match:
                extracted_groups = list(match.groups())
                extracted_groups[1] = '0x' + extracted_groups[1]
                matched_data_list.append(extracted_groups)
    return matched_data_list

def parse_file():
    parse_btn.config(state='disabled')
    try:
        save_config()
        id_index  : int  = letter_to_number(id_place.get())
        data_p   : int  = letter_to_number(data_place.get())
        split_s  : str  = split_entry.get()
        valid_s  : int  = valid_entry.get()
        bms_type : dict = copy.deepcopy(bms_config_backup[protocols_entry.get()]) # type: ignore
        print(
            [id_index       ,
            data_p   ,
            split_s  ,
            valid_s]
        )
        if id_index == '':
            showwarning('警告', '请给出帧ID的')
        elif id_index == -1:
            showwarning('警告', '请给出正确的帧ID位置')
        elif data_p == '':
            showwarning('警告', '请给出帧数据的位置')
        elif data_p == -1:
            showwarning('警告', '请给出正确的帧数据位置')
        elif id_index == data_p:
            showerror('错误', '帧ID和帧数据列数相同')
        elif split_s == '':
            showwarning('警告', "请选择数据分隔符")
        elif valid_s == '':
            showwarning('警告', "请选择有效数据从哪行开始")
        else:
            if ui_path_check(file_path):
                print('正在解析')
                if split_s[0] == 't':
                    split_s = '\t'
                elif '英文逗号' in split_s:
                    split_s = ','
                elif '分号' in split_s:
                    split_s = ';'
                elif '竖线' in split_s:
                    split_s = '|'
                try:
                    file_extension = os.path.splitext(file_path)[1].lower()
                    if file_extension == '.asc':
                        df_data = parse_asc_file(file_path)
                        id_index = 2
                        data_p = 4
                    elif file_extension in ['.xls', '.xlsx']:
                        df_data = read_excel(file_path, int(valid_s))
                    else:
                        df_data = read_csv(file_path, split_s, int(valid_s))
                except Exception as e:
                    print(e)
                    showwarning('警告', f'请先关闭 {os.path.split(file_path)[1]} 文件，我才能工作')
                    return
                if check_data(df_data[3], id_index, data_p):
                    df_data = main_prase(df_data, id_index, data_p, bms_type)
                    output_file_extension = os.path.splitext(file_path)[1].lower()
                    output_base_name = os.path.splitext(os.path.split(file_path)[1])[0]
                    output_file_name = f"{output_base_name}-译{output_file_extension}"

                    try:
                        if output_file_extension in ['.xls', '.xlsx']:
                            creat_excel(file_path, df_data)
                        else:
                            creat_csv(file_path, df_data, split_s)
                    except Exception as e:
                        print(e)
                        showwarning('警告', f'请先关闭 {output_file_name} 文件')
    finally:
        parse_btn.config(state='normal')


def create_sendto_shortcut_ui():
    try:
        sendto_path = os.path.join(os.environ['APPDATA'], 'Microsoft', 'Windows', 'SendTo')
        
        if getattr(sys, 'frozen', False):
            # 如果是PyInstaller打包的exe
            exe_path = os.path.abspath(sys.executable)
            target_path = exe_path
            description = "BMS Translator Application"
        else:
            # 如果是直接运行的Python脚本 (开发模式)
            # 假设ui.py是入口，快捷方式指向python解释器和ui.py
            python_exe = sys.executable
            script_path = os.path.abspath(sys.argv[0])
            target_path = python_exe
            arguments = f'"{script_path}"'
            description = "BMS Translator (Python Script)"

        shortcut_name = "bms-translator.lnk"
        shortcut_path = os.path.join(sendto_path, shortcut_name)

        shell = win32com.client.Dispatch("WScript.Shell")
        shortcut = shell.CreateShortcut(shortcut_path)
        
        shortcut.TargetPath = target_path
        if 'arguments' in locals():
            shortcut.Arguments = arguments
        else:
            shortcut.Arguments = ""
        
        shortcut.WindowStyle = 1 # Normal window
        shortcut.HotKey = ""
        shortcut.IconLocation = exe_path if getattr(sys, 'frozen', False) else os.path.abspath('./image/favicon256x256.ico') # 使用exe的图标或默认图标
        shortcut.Description = description
        shortcut.Save()

        showinfo('成功', f"快捷方式 '{shortcut_name}' 已成功创建在 '{sendto_path}'。")

    except Exception as e:
        showerror('错误', f"创建快捷方式时发生错误: {e}\n请确保您已安装 pywin32 库 (pip install pywin32)。")


def creat_btn():
    file_btn = Button(root, text='选择文件', command=open_file_manager)
    file_btn.place(relx=0.25, y=200)

    global parse_btn
    parse_btn = Button(root, text='开始解析', command=parse_file)
    parse_btn.place(relx=0.55, y=200)

    # 新增的“创建发送到快捷方式”按钮
    sendto_btn = Button(root, text='创建“发送到”', command=create_sendto_shortcut_ui)
    sendto_btn.place(relx=0.05, rely=0.85) # 调整位置避免重叠


def read_config(path):
    if os.path.exists(path):
        with open(path, "r", encoding='utf-8') as fp:
            data = json.load(fp)
        return (data)
    else:
        config = {
            "timestamp": 1,
            "id_place": 5,
            "data_place": 9,
            "split": '英文逗号(,)',
            "valid_row": 2,
            "protocols_type": "GB2015" # 添加默认值，避免KeyError
        }
        with open(path, 'w', encoding='utf-8') as fp:
            json.dump(config, fp)
        return config

def set_config(config):
    if int(time.time()) - config['timestamp'] < 50000: 
        print(config)
        id_place.set(config['id_place'])
        data_place.set(config['data_place'])
        split_entry.set(config['split'])
        valid_entry.set(config['valid_row'])
        protocols_entry.set(config['protocols_type'])
    else:
        id_place.set('5 | E')
        data_place.set("9 | I")
        split_entry.set('英文逗号(,)')
        valid_entry.set(2)
        protocols_entry.set("GB2015")

def letter_to_number(line: str):
    try:
        # 尝试直接将输入转换为整数，处理 "1", "2" 等情况
        return int(line.split(' | ')[0])
    except ValueError:
        # 如果不是纯数字，则尝试按字母处理
        letter = line[0]
        if letter.isalpha():
            # 将字母转换为对应的列数 (A=1, B=2, ...)
            return ord(letter.upper()) - ord('A') + 1
    return -1  # 如果输入不是有效的数字或字母，返回 -1


def not_config_path():
    if os.path.exists('./config'):
        return 1
    else:
        showerror('错误', '没有配置文件夹，请检查')
        root.destroy()

if __name__ == "__main__":
    # 获取当前脚本的绝对路径
    current_script_path = os.path.abspath(sys.argv[0])
    # 获取脚本所在的目录
    current_script_dir = os.path.dirname(current_script_path)
    # 更改当前工作目录到脚本所在的目录
    os.chdir(current_script_dir)

    bms_config = {}
    root = Tk()
    root.resizable(False, False)
    not_config_path()
    if os.path.exists('./config/bms.ico'):
        root.iconbitmap('./config/bms.ico')
    bms_config = read_bms_config('./config/bmsConfig.yaml') # type: ignore
    bms_config_backup = bms_config.copy()
    config = read_config('./config/config')
    win_width, win_height = creat_window()
    id_place, data_place, split_entry, valid_entry, file_path_entry, protocols_entry = creat_entry()

    set_config(config)
    creat_btn()
    
    bms_config = read_bms_config('./config/bmsConfig.yaml') # type: ignore

    # 检查命令行参数
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
        file_path_entry.delete(0, END)
        file_path_entry.insert(0, file_path)
        # parse_file() # 自动解析文件
    
    root.mainloop()
 